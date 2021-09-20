# Load the Pandas libraries with alias 'pd'
import os
from datetime import datetime
from datetime import timedelta

import pandas as pd
import sxl
# importing the modules
from tabulate import tabulate

words_dict = {
    "Total": "الكلي",
    "Service / Specialty": "الخدمة / التخصص",
    "female": "الاناث",
    "male": "الذكور",
    "volunteer": "متطوع",
    "staff": "عامل",
    "service": "الخدمة",
    "program": "البرنامج",
    "category": "البرنامج",

}

words_program_dict = {
    "ambulanceAndEmergency": "الاسعاف والطوارئ",
    "rehab": "التاهيل",
}


# words_col_dict = {
#     "ambulanceAndEmergency": "الاسعاف والطوارئ" ,
#     "rehab":"التاهيل",
# }


class Record:
    def __init__(self):
        self.sector = ""
        self.position = ""
        self.gender = ""
        self.total = 0
        self.date = None
        self.category = ""
        self.fileName = ""

    def to_dict(self):
        if (self.date):
            return dict(Sector=self.sector, Position=self.position, Gender=self.gender, Total=self.total,
                        date=self.date, category=self.category, file=self.fileName)
        return dict(Sector=self.sector, Position=self.position, Gender=self.gender, Total=self.total,
                    category=self.category, file=self.fileName)

    def __repr__(self) -> str:
        return str(self.to_dict())


def getFileName(filePath):
    base = os.path.basename(filePath)
    fileName = os.path.splitext(base)[0]
    return fileName


def getCSVHeaders(filePath, sheetName):
    wb = sxl.Workbook(filePath)
    ws = wb.sheets.get(sheetName)
    return CSVNumberOfStaffAndVolunteersSheetArabic.none_replace(ws.head(3))


class CSVNumberOfStaffAndVolunteersSheetArabic:
    def __init__(self, filePath, sheetName, addDate=False, info=False, offset=0, isSummary=False, df=None,
                 summary_create_path=None):
        if isSummary == False:
            self.df = self.CSV2DataFrame(filePath, sheetName, sheetName, addDate, info, offset=offset)
        else:
            self.DataFrame2CSV(original=filePath, target=summary_create_path, category=sheetName, sheetName=sheetName,
                               df=df)

    @staticmethod
    def CSV2DataFrame(filePath, category, sheetName, addDate=False, info=False, offset=0):
        uploadTime = None
        if (addDate):
            uploadTime = (datetime.today() + timedelta(days=offset)).replace(microsecond=0).isoformat()
        fileName = getFileName(filePath)
        fileName = words_program_dict[fileName] if fileName in words_program_dict else fileName
        # get headers
        headers = getCSVHeaders(filePath, sheetName)

        df = pd.read_excel(filePath, sheet_name=sheetName, skiprows=[0, 1, 2], header=None, engine='openpyxl')
        df.columns = headers[2]
        if (info):
            print(tabulate(df, headers='keys', tablefmt='psql'))
        df_final = pd.DataFrame()
        for index, row in df.iterrows():
            if words_dict['Total'] in row[0].lower():
                continue
            section = -1
            records_l = []
            sections = list(set(headers[1][1:-1]))
            startedColumn = sections[1][0]
            sectorColumn = words_dict["Service / Specialty"]
            for index, column in enumerate(df.columns):

                if headers[1][index] in sections and headers[1][index] != startedColumn:
                    startedColumn = headers[1][index]
                    section += 1
                if (not words_dict['Total'] in column.lower()):
                    if words_dict["female"] in column.lower() or words_dict["male"] in column.lower():
                        record = Record(filePath)
                        record.position = words_dict["volunteer"] if words_dict["volunteer"] in headers[1][
                            index].lower() else words_dict["staff"]
                        record.sector = row[sectorColumn]
                        record.gender = words_dict["female"] if words_dict["female"] in column.lower() else words_dict[
                            "male"]
                        record.total = row[column][section]
                        if (uploadTime):
                            record.date = uploadTime
                        record.category = category
                        record.fileName = fileName
                        records_l.append(record)
                        df_final = df_final.append(record.to_dict(), ignore_index=True)
                    if words_dict["service"] in headers[1][index].lower():
                        sectorColumn = column

        if (info):
            print(tabulate(df_final, headers='keys', tablefmt='psql'))

        return df_final

    @staticmethod
    def none_replace(ls):
        res = ls[:]
        for r_i in range(len(ls), 0, -1):
            for c_i in range(len(ls[r_i - 1])):  # skipp last column which is about total
                cellVal = res[r_i - 1][c_i - 1]
                leftCellVal = res[r_i - 1][c_i - 2] if c_i > 1 else None
                upCellVal = res[r_i - 2][c_i - 1] if r_i > 1 else None
                res[r_i - 1][c_i - 1] = cellVal if cellVal is not None else upCellVal if upCellVal else leftCellVal
        return ls

    @staticmethod
    def DataFrame2CSV(original, target, category, sheetName, df):
        # print(tabulate(df, headers='keys', tablefmt='psql'))
        # print(original, target, sheetName)
        # shutil.copyfile(original, target)
        import openpyxl
        import pandas as pd
        # split and merge by Gender
        dfs = [rows for _, rows in df.groupby('Gender')]
        combined = pd.merge(dfs[0].drop(['date'], axis=1), dfs[1].drop(['date'], axis=1),
                            on=["category", "file", "Position", ])

        # split and merge by Position
        dfs = [rows for _, rows in combined.groupby('Position')]
        # print(tabulate(dfs[0], headers='keys', tablefmt='psql'))

        # combined = pd.merge(dfs[0], dfs[1], on=["category","file"])

        filter_gender_col1 = [col for col in dfs[0] if col.startswith('Gender')]
        filter_gender_col2 = [col for col in dfs[1] if col.startswith('Gender')]
        filter_gender_col = [filter_gender_col1, filter_gender_col2]
        for i in range(len(dfs)):
            columns_dict_rename = dict()
            for gender_col in filter_gender_col[i]:
                columns_dict_rename["count" + gender_col[gender_col.find("_"):]] = dfs[i][gender_col].iloc[0]
            dfs[i].rename(columns=columns_dict_rename, inplace=True)
            dfs[i].loc[:, words_dict['Total']] = dfs[i].sum(numeric_only=True, axis=1)
            dfs[i].drop(filter_gender_col[i], axis=1, inplace=True)

            dfs[i]["Position"].replace({"عامل": "عدد العاملين",
                                        "متطوع": "عدد المتطوعين"}, inplace=True)

            # dfs[i]['Position'] = df['Position'].replace(['عامل'], 'عدد العاملين')

            print(tabulate(dfs[i], headers='keys', tablefmt='psql'))

        # Rename columns

        combined_df = pd.concat(dfs)

        def isCategory(columnTitle):
            return isinstance(columnTitle, str) and \
                   columnTitle.lower() in ['category', 'program', words_dict['category'].lower()]

        def isTotal(columnTitle):
            return isinstance(columnTitle, str) and \
                   columnTitle.strip() in [words_dict['Total'], 'Total']

        def isGender(columnTitle):
            print('[', columnTitle, ']')
            return isinstance(columnTitle, str) and \
                   columnTitle.strip() in [words_dict['female'], words_dict['male'], 'male', 'female']

        def isPosition(columnTitle):
            return isinstance(columnTitle, str) and \
                   columnTitle.strip() in ['عدد العاملين', 'عدد المتطوعين']

        def get_column_name(sh, row, column):
            # for i in range(1, row):
            cell_obj = sh.cell(row, column)
            if (cell_obj.value):
                return cell_obj.value
            return None

        # pass
        from pprint import pprint
        wb = openpyxl.load_workbook(target)
        sh = wb.active
        pos = dict()
        positionVisited = None
        totalVisited = None
        genderVisited = dict()
        # iterate through excel and display data
        for j in range(1, sh.max_column + 1):
            for i in range(1, sh.max_row + 1):
                titleVal = get_column_name(sh, i, j)
                if titleVal and isCategory(titleVal):
                    for row_index in range(i + 1, sh.max_row + 1):
                        cellVal = get_column_name(sh, row_index, j)
                        if cellVal and cellVal != 'الاجمالي':
                            pos[row_index] = {'program': cellVal}
                    break

                if titleVal and isPosition(titleVal):
                    positionVisited = titleVal

                if titleVal and isGender(titleVal):
                    genderVisited.update({titleVal: j})
                    break

                if titleVal and isTotal(titleVal):
                    totalVisited = {titleVal: j}
                    break

                if positionVisited and totalVisited and genderVisited and len(genderVisited) == 2:
                    totalVisited.update(genderVisited)
                    for row_index in pos:
                        pos[row_index][positionVisited] = totalVisited
                    positionVisited = None
                    totalVisited = None
                    genderVisited = dict()
                    break

        pprint(pos)

        position, desired_columns = [], []
        program = None
        for row_index in pos:
            iteration_times = len(pos[row_index])
            for i, key in enumerate(pos[row_index]):
                if isPosition(key):
                    position.append(key)
                    desired_columns.append(pos[row_index][key])
                elif isCategory(key):
                    program = pos[row_index][key]
                if position and program and i == iteration_times - 1:
                    print(position, program, desired_columns)
                    position, desired_columns = [], []
                    program = None

        # cell_obj = sh.cell(row=i, column=j)
        # print(cell_obj.value, end=" ")
        # sh.cell(row=5, column=5).value = 2
        wb.save(target)
